from __future__ import annotations

import json
import hashlib
from pathlib import Path
from datetime import date, datetime, timedelta

import pandas as pd
import openpyxl
import psycopg2
import psycopg2.extras
import streamlit as st

BASE_DIR = Path(__file__).resolve().parents[1]
DATA_DIR = BASE_DIR / "data"
DEFAULT_XLSM = BASE_DIR / "assets" / "LMP.xlsm"

DEFAULT_PERMISSIONS = {
    "view_dashboard": True,
    "view_alerts": True,
    "view_employee_db": True,
    "view_employee_panel": True,
    "create_referrals": True,
    "view_referrals": True,
    "view_hazard_map": True,
    "import_data": True,
    "export_data": True,
    "manage_users": True,
    "security_center": True,
    "history": True,
}

ROLE_PRESETS = {
    "Administrator": DEFAULT_PERMISSIONS,
    "BHP": {**DEFAULT_PERMISSIONS, "manage_users": False},
    "HR": {
        "view_dashboard": True,
        "view_alerts": True,
        "view_employee_db": True,
        "view_employee_panel": True,
        "create_referrals": True,
        "view_referrals": True,
        "view_hazard_map": True,
        "import_data": False,
        "export_data": True,
        "manage_users": False,
        "security_center": False,
        "history": False,
    },
    "Podgląd": {
        "view_dashboard": True,
        "view_alerts": True,
        "view_employee_db": True,
        "view_employee_panel": True,
        "create_referrals": False,
        "view_referrals": True,
        "view_hazard_map": True,
        "import_data": False,
        "export_data": False,
        "manage_users": False,
        "security_center": False,
        "history": False,
    },
}

CATEGORY_TO_SECTION = {
    "CZYNNIKI FIZYCZNE": "I. Czynniki fizyczne",
    "PYŁY": "II. Pyły",
    "CZYNNIKI CHEMICZNE": "III. Czynniki chemiczne",
    "CZYNNIKI BIOLOGICZNE": "IV. Czynniki biologiczne",
    "INNE": "V. Inne czynniki, w tym niebezpieczne",
}
SECTION_OPTIONS = list(CATEGORY_TO_SECTION.values())


def get_connection():
    return psycopg2.connect(
        st.secrets["DB_URL"],
        cursor_factory=psycopg2.extras.RealDictCursor,
    )


def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode("utf-8")).hexdigest()


def init_db() -> None:
    conn = get_connection()
    cur = conn.cursor()

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            login TEXT UNIQUE NOT NULL,
            full_name TEXT NOT NULL,
            role TEXT NOT NULL,
            password_hash TEXT NOT NULL,
            active BOOLEAN NOT NULL DEFAULT TRUE,
            failed_attempts INTEGER NOT NULL DEFAULT 0,
            blocked_until TIMESTAMP NULL,
            last_login TIMESTAMP NULL,
            permissions_json TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS departments (
            id SERIAL PRIMARY KEY,
            name TEXT UNIQUE NOT NULL
        );

        CREATE TABLE IF NOT EXISTS positions (
            id SERIAL PRIMARY KEY,
            department_id INTEGER NOT NULL REFERENCES departments(id) ON DELETE CASCADE,
            name TEXT NOT NULL,
            UNIQUE(department_id, name)
        );

        CREATE TABLE IF NOT EXISTS hazard_map (
            id SERIAL PRIMARY KEY,
            department_name TEXT NOT NULL,
            position_name TEXT NOT NULL,
            hazard_name TEXT NOT NULL,
            category TEXT NOT NULL,
            section_label TEXT NOT NULL,
            work_conditions TEXT
        );

        CREATE TABLE IF NOT EXISTS employees (
            id SERIAL PRIMARY KEY,
            full_name TEXT NOT NULL,
            department_name TEXT NOT NULL,
            position_name TEXT NOT NULL,
            pesel TEXT,
            address TEXT,
            last_exam_date DATE NULL,
            next_exam_date DATE NULL,
            status TEXT,
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS referrals (
            id SERIAL PRIMARY KEY,
            referral_number TEXT UNIQUE NOT NULL,
            employee_id INTEGER NULL REFERENCES employees(id) ON DELETE SET NULL,
            employee_name TEXT NOT NULL,
            department_name TEXT NOT NULL,
            position_name TEXT NOT NULL,
            position_description TEXT,
            issue_date DATE NOT NULL,
            next_exam_date DATE NULL,
            exam_type TEXT NOT NULL,
            employer TEXT,
            pesel TEXT,
            employee_address TEXT,
            place_of_issue TEXT,
            status TEXT,
            work_conditions TEXT,
            hazards_count INTEGER NOT NULL DEFAULT 0,
            pdf_path TEXT,
            created_by TEXT,
            created_at TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP
        );

        CREATE TABLE IF NOT EXISTS referral_hazards (
            id SERIAL PRIMARY KEY,
            referral_id INTEGER NOT NULL REFERENCES referrals(id) ON DELETE CASCADE,
            hazard_name TEXT NOT NULL,
            category TEXT NOT NULL,
            section_label TEXT NOT NULL,
            work_conditions TEXT
        );

        CREATE TABLE IF NOT EXISTS audit_log (
            id SERIAL PRIMARY KEY,
            event_time TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
            username TEXT,
            action TEXT NOT NULL,
            details TEXT
        );
        """
    )
    conn.commit()

    cur.execute("SELECT COUNT(*) AS cnt FROM users")
    if cur.fetchone()["cnt"] == 0:
        seed_users = [
            ("admin", "Administrator BHP", "Administrator", "Admin123!@#"),
            ("bhp", "Specjalista BHP", "BHP", "Bhp123!@#45"),
            ("hr", "Specjalista HR", "HR", "Hr123!@#45"),
            ("podglad", "Podgląd", "Podgląd", "Podglad123!@#"),
        ]
        for login, full_name, role, password in seed_users:
            perms = json.dumps(ROLE_PRESETS[role], ensure_ascii=False)
            cur.execute(
                """
                INSERT INTO users (login, full_name, role, password_hash, permissions_json)
                VALUES (%s, %s, %s, %s, %s)
                """,
                (login, full_name, role, hash_password(password), perms),
            )
        conn.commit()

    cur.execute("SELECT COUNT(*) AS cnt FROM hazard_map")
    if cur.fetchone()["cnt"] == 0 and DEFAULT_XLSM.exists():
        import_hazard_map(DEFAULT_XLSM, replace=True, conn=conn)

    cur.execute("SELECT COUNT(*) AS cnt FROM employees")
    if cur.fetchone()["cnt"] == 0:
        seed_sample_data(conn)

    conn.close()


def import_hazard_map(file_path: str | Path, replace: bool = True, conn=None) -> int:
    own = conn is None
    if own:
        conn = get_connection()

    wb = openpyxl.load_workbook(file_path, data_only=True, keep_vba=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    df = pd.DataFrame(rows[1:], columns=[str(x).strip() if x else "" for x in rows[0]])
    df = df.rename(
        columns={
            "DZIAŁ": "department",
            "STANOWISKO": "position",
            "ZAGROŻENIA": "hazard",
            "KATEGORIA": "category",
            "OPIS WARUNKÓW PRACY": "work_conditions",
        }
    )

    needed = ["department", "position", "hazard", "category", "work_conditions"]
    for col in needed:
        if col not in df.columns:
            df[col] = ""

    df = df[needed].dropna(subset=["department", "position", "hazard"])
    df["category"] = df["category"].fillna("INNE").astype(str).str.strip().replace({"": "INNE"})

    def normalize_section(cat: str) -> str:
        cat = str(cat).strip().upper()
        return CATEGORY_TO_SECTION.get(cat, "V. Inne czynniki, w tym niebezpieczne")

    df["section_label"] = df["category"].apply(normalize_section)

    cur = conn.cursor()

    if replace:
        cur.execute("DELETE FROM referral_hazards")
        cur.execute("DELETE FROM hazard_map")
        cur.execute("DELETE FROM positions")
        cur.execute("DELETE FROM departments")

    departments = sorted(df["department"].dropna().astype(str).str.strip().unique())
    for dep in departments:
        cur.execute(
            """
            INSERT INTO departments(name)
            VALUES (%s)
            ON CONFLICT (name) DO NOTHING
            """,
            (dep,),
        )
    conn.commit()

    for _, row in df.iterrows():
        dep = str(row["department"]).strip()
        pos = str(row["position"]).strip()

        cur.execute("SELECT id FROM departments WHERE name = %s", (dep,))
        dep_row = cur.fetchone()
        if not dep_row:
            continue
        dep_id = dep_row["id"]

        cur.execute(
            """
            INSERT INTO positions(department_id, name)
            VALUES (%s, %s)
            ON CONFLICT (department_id, name) DO NOTHING
            """,
            (dep_id, pos),
        )

        cur.execute(
            """
            INSERT INTO hazard_map (
                department_name, position_name, hazard_name, category, section_label, work_conditions
            )
            VALUES (%s, %s, %s, %s, %s, %s)
            """,
            (
                dep,
                pos,
                str(row["hazard"]).strip(),
                str(row["category"]).strip().upper(),
                row["section_label"],
                str(row["work_conditions"] or "").strip(),
            ),
        )

    conn.commit()
    count = len(df)
    log_action("system", "IMPORT_MAPA_ZAGROZEN", f"Zaimportowano {count} rekordów", conn)

    if own:
        conn.close()

    return count


def seed_sample_data(conn=None) -> None:
    own = conn is None
    if own:
        conn = get_connection()

    cur = conn.cursor()
    today = date.today()

    samples = [
        (
            "Adam Nowak",
            "DZIAŁ INWESTYCJI",
            "Dyrektor/Zastępca dyrektora/Dyrektor projektu",
            "",
            "",
            today - timedelta(days=2),
            today - timedelta(days=2),
            "PO TERMINIE",
        ),
        (
            "Andrzej Kowalski",
            "DR",
            "Specjalista/Główny specjalista/Starszy specjalista",
            "",
            "",
            today - timedelta(days=7),
            today + timedelta(days=7),
            "OK",
        ),
        (
            "Janusz Kowalski",
            "DZIAŁ HANDLOWY",
            "Kasjer",
            "",
            "",
            today - timedelta(days=30),
            today + timedelta(days=20),
            "KOŃCZY SIĘ W 30 DNI",
        ),
        (
            "Pracownik z arkusza zagrożeń",
            "DU WSO",
            "Mistrz",
            "",
            "",
            None,
            None,
            "BRAK BADAŃ",
        ),
        (
            "Roman Janusz",
            "DZIAŁ INWESTYCJI",
            "Dyrektor/Zastępca dyrektora/Dyrektor projektu",
            "",
            "",
            today - timedelta(days=10),
            today - timedelta(days=1),
            "PO TERMINIE",
        ),
    ]

    for s in samples:
        cur.execute(
            """
            INSERT INTO employees (
                full_name, department_name, position_name, pesel, address,
                last_exam_date, next_exam_date, status
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """,
            s,
        )

    conn.commit()

    if own:
        conn.close()


def log_action(username: str, action: str, details: str = "", conn=None):
    own = conn is None
    if own:
        conn = get_connection()

    cur = conn.cursor()
    cur.execute(
        """
        INSERT INTO audit_log (username, action, details)
        VALUES (%s, %s, %s)
        """,
        (username, action, details),
    )
    conn.commit()

    if own:
        conn.close()


def authenticate(login: str, password: str):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM users WHERE login = %s", (login,))
    user = cur.fetchone()

    if not user:
        conn.close()
        return None, "Nieprawidłowy login lub hasło."

    if not user["active"]:
        conn.close()
        return None, "Konto jest zablokowane przez administratora."

    if user["blocked_until"]:
        blocked_until = user["blocked_until"]
        if blocked_until > datetime.now():
            conn.close()
            return None, f"Konto czasowo zablokowane do {blocked_until:%Y-%m-%d %H:%M}."

    if hash_password(password) != user["password_hash"]:
        failed = user["failed_attempts"] + 1
        blocked_until = None

        if failed >= 5:
            blocked_until = datetime.now() + timedelta(minutes=15)
            failed = 0

        cur.execute(
            """
            UPDATE users
            SET failed_attempts = %s, blocked_until = %s
            WHERE id = %s
            """,
            (failed, blocked_until, user["id"]),
        )
        conn.commit()
        conn.close()
        return None, "Nieprawidłowy login lub hasło."

    cur.execute(
        """
        UPDATE users
        SET failed_attempts = 0, blocked_until = NULL, last_login = %s
        WHERE id = %s
        """,
        (datetime.now(), user["id"]),
    )
    conn.commit()

    result = dict(user)
    result["permissions"] = json.loads(user["permissions_json"])
    conn.close()
    return result, None


def get_departments() -> list[str]:
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT name FROM departments ORDER BY name")
    rows = cur.fetchall()
    conn.close()
    return [r["name"] for r in rows]


def get_positions(department_name: str | None = None) -> list[str]:
    conn = get_connection()
    cur = conn.cursor()

    if department_name and department_name != "Wszystkie":
        cur.execute(
            """
            SELECT DISTINCT position_name
            FROM hazard_map
            WHERE department_name = %s
            ORDER BY position_name
            """,
            (department_name,),
        )
    else:
        cur.execute(
            """
            SELECT DISTINCT position_name
            FROM hazard_map
            ORDER BY position_name
            """
        )

    rows = cur.fetchall()
    conn.close()
    return [r["position_name"] for r in rows]


def get_hazards(department_name: str, position_name: str) -> pd.DataFrame:
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        """
        SELECT department_name, position_name, hazard_name, category, section_label, work_conditions
        FROM hazard_map
        WHERE department_name = %s AND position_name = %s
        ORDER BY hazard_name
        """,
        (department_name, position_name),
    )
    rows = cur.fetchall()
    conn.close()

    return pd.DataFrame(
        rows,
        columns=["Dział", "Stanowisko", "Zagrożenie", "Kategoria", "Sekcja", "Opis warunków pracy"],
    )


def get_employees_df():
    conn = get_connection()
    df = pd.read_sql_query("SELECT * FROM employees ORDER BY full_name", conn)
    conn.close()

    if df.empty:
        return df

    today = date.today()

    def parse_date(value):
        if value is None or value == "":
            return None
        if isinstance(value, str):
            for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
                try:
                    return datetime.strptime(value, fmt).date()
                except Exception:
                    pass
            try:
                return datetime.fromisoformat(value).date()
            except Exception:
                return None
        if hasattr(value, "date"):
            try:
                return value.date()
            except Exception:
                pass
        if isinstance(value, date):
            return value
        return None

    df["last_exam_date_parsed"] = df["last_exam_date"].apply(parse_date) if "last_exam_date" in df.columns else None
    df["next_exam_date_parsed"] = df["next_exam_date"].apply(parse_date) if "next_exam_date" in df.columns else None

    def calc_days(d):
        if d is None:
            return None
        return (d - today).days

    df["dni_do_badan"] = df["next_exam_date_parsed"].apply(calc_days)

    def calc_status(row):
        next_date = row["next_exam_date_parsed"]
        days = row["dni_do_badan"]

        if next_date is None:
            if row.get("last_exam_date_parsed", None) is None:
                return "BRAK BADAŃ"
            return "BRAK DATY"

        if days is None:
            return "BRAK DATY"
        if days < 0:
            return "PO TERMINIE"
        if days <= 30:
            return "KOŃCZY SIĘ W 30 DNI"
        return "OK"

    df["status"] = df.apply(calc_status, axis=1)

    df = df.drop(columns=["last_exam_date_parsed", "next_exam_date_parsed"], errors="ignore")
    return df


def days_to_exam(next_exam_date: str | None):
    if not next_exam_date:
        return None
    try:
        return (date.fromisoformat(next_exam_date) - date.today()).days
    except Exception:
        return None


def compute_status(next_exam_date):
    if not next_exam_date:
        return "BRAK DATY"

    if isinstance(next_exam_date, date):
        d = (next_exam_date - date.today()).days
    else:
        d = days_to_exam(str(next_exam_date))

    if d is None:
        return "BRAK DATY"
    if d < 0:
        return "PO TERMINIE"
    if d <= 30:
        return "KOŃCZY SIĘ W 30 DNI"
    return "OK"


def upsert_employee(employee: dict, conn=None) -> int:
    own = conn is None
    if own:
        conn = get_connection()

    cur = conn.cursor()
    status = compute_status(employee.get("next_exam_date"))

    if employee.get("id"):
        cur.execute(
            """
            UPDATE employees
            SET
                full_name = %s,
                department_name = %s,
                position_name = %s,
                pesel = %s,
                address = %s,
                last_exam_date = %s,
                next_exam_date = %s,
                status = %s
            WHERE id = %s
            """,
            (
                employee["full_name"],
                employee["department_name"],
                employee["position_name"],
                employee.get("pesel") or "",
                employee.get("address") or "",
                employee.get("last_exam_date"),
                employee.get("next_exam_date"),
                status,
                employee["id"],
            ),
        )
        emp_id = employee["id"]
    else:
        cur.execute(
            """
            INSERT INTO employees (
                full_name, department_name, position_name, pesel, address,
                last_exam_date, next_exam_date, status
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id
            """,
            (
                employee["full_name"],
                employee["department_name"],
                employee["position_name"],
                employee.get("pesel") or "",
                employee.get("address") or "",
                employee.get("last_exam_date"),
                employee.get("next_exam_date"),
                status,
            ),
        )
        emp_id = cur.fetchone()["id"]

    conn.commit()

    if own:
        conn.close()

    return emp_id


def next_referral_number(issue_date: date | str, conn=None) -> str:
    own = conn is None
    if own:
        conn = get_connection()

    d = date.fromisoformat(issue_date) if isinstance(issue_date, str) else issue_date
    prefix = f"/{d.month:02d}/{d.year}"

    cur = conn.cursor()
    cur.execute(
        """
        SELECT COUNT(*) AS cnt
        FROM referrals
        WHERE EXTRACT(MONTH FROM issue_date) = %s
          AND EXTRACT(YEAR FROM issue_date) = %s
        """,
        (d.month, d.year),
    )
    cnt = cur.fetchone()["cnt"]

    if own:
        conn.close()

    return f"{cnt + 1:03d}{prefix}"


def create_referral(payload: dict, hazards: list[dict], username: str) -> int:
    conn = get_connection()
    cur = conn.cursor()

    issue_date = payload["issue_date"]
    referral_number = next_referral_number(issue_date, conn)
    status = compute_status(payload.get("next_exam_date"))

    if payload.get("employee_id"):
        emp_id = payload["employee_id"]
        upsert_employee(
            {
                "id": emp_id,
                "full_name": payload["employee_name"],
                "department_name": payload["department_name"],
                "position_name": payload["position_name"],
                "pesel": payload.get("pesel"),
                "address": payload.get("employee_address"),
                "last_exam_date": issue_date,
                "next_exam_date": payload.get("next_exam_date"),
            },
            conn,
        )
    else:
        emp_id = upsert_employee(
            {
                "full_name": payload["employee_name"],
                "department_name": payload["department_name"],
                "position_name": payload["position_name"],
                "pesel": payload.get("pesel"),
                "address": payload.get("employee_address"),
                "last_exam_date": issue_date,
                "next_exam_date": payload.get("next_exam_date"),
            },
            conn,
        )

    cur.execute(
        """
        INSERT INTO referrals (
            referral_number, employee_id, employee_name, department_name, position_name,
            position_description, issue_date, next_exam_date, exam_type, employer,
            pesel, employee_address, place_of_issue, status, work_conditions,
            hazards_count, created_by
        )
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        RETURNING id
        """,
        (
            referral_number,
            emp_id,
            payload["employee_name"],
            payload["department_name"],
            payload["position_name"],
            payload.get("position_description"),
            payload["issue_date"],
            payload.get("next_exam_date"),
            payload["exam_type"],
            payload.get("employer"),
            payload.get("pesel"),
            payload.get("employee_address"),
            payload.get("place_of_issue"),
            status,
            payload.get("work_conditions"),
            len(hazards),
            username,
        ),
    )
    rid = cur.fetchone()["id"]

    for hz in hazards:
        cur.execute(
            """
            INSERT INTO referral_hazards (referral_id, hazard_name, category, section_label, work_conditions)
            VALUES (%s, %s, %s, %s, %s)
            """,
            (
                rid,
                hz.get("hazard_name", ""),
                hz.get("category", ""),
                hz.get("section_label", ""),
                hz.get("work_conditions", ""),
            ),
        )

    conn.commit()
    log_action(username, "NOWE_SKIEROWANIE", f'{referral_number} {payload["employee_name"]}', conn)
    conn.close()
    return rid


def update_referral_pdf_path(referral_id: int, pdf_path: str):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("UPDATE referrals SET pdf_path = %s WHERE id = %s", (pdf_path, referral_id))
    conn.commit()
    conn.close()


def get_referral(referral_id: int) -> dict:
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("SELECT * FROM referrals WHERE id = %s", (referral_id,))
    referral = cur.fetchone()
    if not referral:
        conn.close()
        return {}

    cur.execute(
        """
        SELECT hazard_name, category, section_label, work_conditions
        FROM referral_hazards
        WHERE referral_id = %s
        ORDER BY id
        """,
        (referral_id,),
    )
    hazards = [dict(r) for r in cur.fetchall()]

    conn.close()
    referral = dict(referral)
    referral["hazards"] = hazards
    return referral


def get_referrals_df() -> pd.DataFrame:
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM referrals ORDER BY issue_date DESC, id DESC")
    rows = cur.fetchall()
    conn.close()
    return pd.DataFrame(rows)


def get_users_df() -> pd.DataFrame:
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM users ORDER BY role, login")
    rows = cur.fetchall()
    conn.close()
    return pd.DataFrame(rows)


def create_user(login: str, full_name: str, role: str, password: str):
    conn = get_connection()
    cur = conn.cursor()
    perms = json.dumps(ROLE_PRESETS.get(role, DEFAULT_PERMISSIONS), ensure_ascii=False)
    cur.execute(
        """
        INSERT INTO users(login, full_name, role, password_hash, permissions_json)
        VALUES (%s, %s, %s, %s, %s)
        """,
        (login, full_name, role, hash_password(password), perms),
    )
    conn.commit()
    conn.close()


def set_user_active(user_id: int, active: bool):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("UPDATE users SET active = %s WHERE id = %s", (active, user_id))
    conn.commit()
    conn.close()


def unlock_user(user_id: int):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        "UPDATE users SET blocked_until = NULL, failed_attempts = 0 WHERE id = %s",
        (user_id,),
    )
    conn.commit()
    conn.close()


def set_user_permissions(user_id: int, permissions: dict):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        "UPDATE users SET permissions_json = %s WHERE id = %s",
        (json.dumps(permissions, ensure_ascii=False), user_id),
    )
    conn.commit()
    conn.close()


def reset_user_password(user_id: int, new_password: str):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute(
        "UPDATE users SET password_hash = %s WHERE id = %s",
        (hash_password(new_password), user_id),
    )
    conn.commit()
    conn.close()


def get_audit_df() -> pd.DataFrame:
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM audit_log ORDER BY id DESC")
    rows = cur.fetchall()
    conn.close()
    return pd.DataFrame(rows)


def update_employee(
    employee_id: int,
    full_name: str,
    department_name: str,
    position_name: str,
    pesel: str,
    address: str,
    last_exam_date: str | None,
    next_exam_date: str | None,
):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute(
        """
        UPDATE employees
        SET
            full_name = %s,
            department_name = %s,
            position_name = %s,
            pesel = %s,
            address = %s,
            last_exam_date = %s,
            next_exam_date = %s,
            status = %s
        WHERE id = %s
        """,
        (
            full_name,
            department_name,
            position_name,
            pesel,
            address,
            last_exam_date,
            next_exam_date,
            compute_status(next_exam_date),
            employee_id,
        ),
    )

    conn.commit()
    conn.close()


def delete_employee(employee_id: int):
    conn = get_connection()
    cur = conn.cursor()

    cur.execute("DELETE FROM referrals WHERE employee_id = %s", (employee_id,))
    cur.execute("DELETE FROM employees WHERE id = %s", (employee_id,))

    conn.commit()
    conn.close()